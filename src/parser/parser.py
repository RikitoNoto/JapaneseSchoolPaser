from abc import ABC, abstractmethod
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from typing import Optional


class Parser(ABC):
    def __init__(self, sheet: Worksheet) -> None:
        self._sheet = sheet

    @abstractmethod
    def parse():
        pass

    def _search_cell(self, keyword: str, sheet: Worksheet) -> Optional[Cell]:
        """
        keywordをシートから検索し、最初に見つけたセルを返す。
        検索はA1→A2→B2→B1→A3→B3→C3→C2→C1
        のようにA1から(max_row, max_column)に直線を引くような方向で検索をする。

        Args:
            keyword (str): 検索するキーワード
            sheet (Worksheet): 検索対象のシート
        """
        # 行か列の大きい方でループ
        for i in range(max([sheet.max_column, sheet.max_row])):
            column = i + 1
            # columnを縦に検索
            for j in range(column - 1):
                row = i + 1
                column = j + 1
                if sheet.cell(row=row, column=column).value == keyword:
                    return sheet.cell(row=row, column=column)

            # rowとcolumnが同じセルを検索
            if sheet.cell(row=column, column=column).value == keyword:
                return sheet.cell(row=column, column=column)

            # rowを右から左に検索
            for j in reversed(range(column - 1)):
                row = j + 1
                if sheet.cell(row=row, column=column).value == keyword:
                    return sheet.cell(row=row, column=column)
