from typing import Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from src.models.graduate_school import GraduateSchool, Major
from src.parser.parser import Parser


class GraduateSchoolParser(Parser):
    def parse(self) -> list[GraduateSchool]:
        """
        研究科のセルを基準にデータを検索しパースする。
        """
        base_cell: Optional[Cell] = self._search_cell(
            "研究科",
            self._sheet,
        )
        if not base_cell:
            return []

        row = base_cell.row + 3  # 研究科の3行下から開始
        column = base_cell.column
        graduate_schools: list[GraduateSchool] = []
        # 空白のセルが見つかるまで、下を検索
        while self._sheet.cell(row=row, column=column).value not in [None, ""]:
            graduate_school_name = self._sheet.cell(row=row, column=column).value

            # 初めて見つけた学科の場合は新規追加
            if graduate_school_name not in [g.name for g in graduate_schools]:
                graduate_schools.append(
                    GraduateSchool(
                        self._sheet.cell(row=row, column=column).value,
                        [
                            Major(
                                self._sheet.cell(row=row, column=column + 2).value,
                            )
                        ],
                    )
                )
            else:
                pass
                for graduate_school in graduate_schools:
                    if graduate_school.name == graduate_school_name:
                        graduate_school.majors.append(
                            Major(self._sheet.cell(row=row, column=column + 2).value),
                        ),
            row += 1

        return graduate_schools
