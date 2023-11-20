import re

from typing import Optional
from openpyxl.cell.cell import Cell
from models.base_info import BaseInfo, SchoolClassification
from parser.parser import Parser


class BaseInfoParser(Parser):
    TITLE_CELL: dict[str, int] = {"row": 1, "column": 2}

    def parse(self) -> BaseInfo:
        """
        学校コードのセルを基準にデータを検索しパースする。
        それに加えて、ファイル名とB1セルから大学名と学校タイプをパースする。
        学校コードが見つからない場合は、情報が取得できないためValueErrorをraiseする。
        """
        base_cell: Optional[Cell] = self._search_cell(
            "学校コード",
            self._sheet,
        )
        if not base_cell:
            raise ValueError("学校コードが見つかりませんでした。")

        name = self._sheet.title  # シート名が大学名

        title = self._sheet.cell(**self.TITLE_CELL).value

        classification: Optional[SchoolClassification] = None
        classification_match: Optional[re.Match[str]] = re.search(
            r"(.*?) ", title
        )  # 最初の空白より前が大学種
        if classification_match:
            classification_str = classification_match.group(1)
            classification = SchoolClassification.from_str(classification_str)

        name_en: str = ""
        name_en_match: Optional[re.Match[str]] = re.search(r"（(.*?)）", title)  # 括弧の中が英名
        if name_en_match:
            name_en = name_en_match.group(1)

        return BaseInfo(
            name=name,
            name_en=name_en,
            school_code=self._sheet.cell(
                row=base_cell.row + 1, column=base_cell.column
            ).value,
            president=self._sheet.cell(
                row=base_cell.row + 1, column=base_cell.column + 1
            ).value,
            classification=classification,
        )
