from typing import Optional
from openpyxl.cell.cell import Cell
from models.faculty import Department, Faculty
from parser.parser import Parser


class FacultyParser(Parser):
    def parse(self) -> list[Faculty]:
        """
        学部のセルを基準にデータを検索しパースする。
        """
        base_cell: Optional[Cell] = self._search_cell(
            "学部",
            self._sheet,
        )
        if not base_cell:
            return []

        row = base_cell.row + 3  # 学部の3行下から開始
        column = base_cell.column
        faculties: list[Faculty] = []
        # 空白のセルが見つかるまで、下を検索
        while self._sheet.cell(row=row, column=column).value not in [None, ""]:
            faculty_name = self._sheet.cell(row=row, column=column).value

            # 初めて見つけた学科の場合は新規追加
            if faculty_name not in [f.name for f in faculties]:
                faculties.append(
                    Faculty(
                        self._sheet.cell(row=row, column=column).value,
                        [
                            Department(
                                self._sheet.cell(row=row, column=column + 2).value,
                            )
                        ],
                    )
                )
            else:
                for faculty in faculties:
                    if faculty.name == faculty_name:
                        faculty.departments.append(
                            Department(
                                self._sheet.cell(row=row, column=column + 2).value
                            ),
                        ),
            row += 1

        return faculties
