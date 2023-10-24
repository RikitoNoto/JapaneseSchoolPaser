import json
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from models.school import School
from parser.school_parser import SchoolParser


def parse_schools_to_dict(path: str) -> dict[str, Any]:
    schools = parse_schools_to_model(path)

    return {"schools": [school.to_dict() for school in schools]}


def parse_schools_to_model(path: str) -> list[School]:
    book: Workbook = load_workbook(path)
    schools: list[School] = []

    sheet_names = book.sheetnames
    for sheet_name in sheet_names:
        sheet: Worksheet = book[sheet_name]
        schools.append(SchoolParser(sheet).parse())

    book.close()

    return schools


def output_json(source_path: str, out_file_path: str) -> None:
    schools = parse_schools_to_dict(source_path)
    with open(out_file_path, "w") as file:
        json.dump(schools, file)


if __name__ == "__main__":
    import sys

    args = sys.argv
    output_json(args[1], args[2])
